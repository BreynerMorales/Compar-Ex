# This Script generate report in txt file and json
import json
import os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from tkinter import Tk, filedialog
from datetime import datetime

def default_converter(o):
    if isinstance(o, datetime):
        return o.isoformat()  # Converts to 'YYYY-MM-DDTHH:MM:SS'
    raise TypeError(f'Object of type {o.__class__.__name__} is not JSON serializable')

def seleccionar_archivo(titulo):
    Tk().withdraw()
    return filedialog.askopenfilename(
        title=titulo,
        filetypes=[("Archivos Excel", "*.xlsx")]
    )

def seleccionar_guardado():
    Tk().withdraw()
    return filedialog.asksaveasfilename(
        title="Guardar reporte de diferencias",
        defaultextension=".txt",
        filetypes=[("Archivo de texto", "*.txt")]
    )

def obtener_encabezados(ws):
    encabezados = {}
    for col in range(1, ws.max_column + 1):
        valor = ws.cell(row=1, column=col).value
        if valor is None:
            encabezados[col] = f"SIN_TITULO_{get_column_letter(col)}"
        else:
            encabezados[col] = str(valor).strip()
    return encabezados

def comparar_xlsx(archivo1, archivo2, archivo_reporte, archivo_json):
    wb1 = load_workbook(archivo1, data_only=True)
    wb2 = load_workbook(archivo2, data_only=True)

    nombre1 = os.path.basename(archivo1)
    nombre2 = os.path.basename(archivo2)
    fecha_ejecucion = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    total_diferencias = 0
    diferencias_por_hoja = {}
    hojas_faltantes = 0
    resumen_json = {
        "meta": {
            "archivo1": nombre1,
            "archivo2": nombre2,
            "fecha_ejecucion": fecha_ejecucion,
            "user_execute": "deister_bmp"  # Opción para agregar usuario que ejecuta
        },
        "resumen": {
            "hojas_comparadas": 0,
            "hojas_faltantes": 0,
            "total_diferencias": 0
        },
        "hojas": []
    }

    with open(archivo_reporte, "w", encoding="utf-8") as reporte:
        # ENCABEZADO
        reporte.write("REPORTE DE DIFERENCIAS ENTRE LIBROS XLSX\n")
        reporte.write("=" * 70 + "\n")
        reporte.write(f"Archivo 1        : {nombre1}\n")
        reporte.write(f"Archivo 2        : {nombre2}\n")
        reporte.write(f"Fecha ejecución  : {fecha_ejecucion}\n")
        reporte.write("\n")

        hojas1 = set(wb1.sheetnames)
        hojas2 = set(wb2.sheetnames)

        # Hojas faltantes
        for hoja in sorted(hojas1 - hojas2):
            reporte.write(f"[HOJA FALTANTE] '{hoja}' solo existe en ARCHIVO 1\n")
            hojas_faltantes += 1

        for hoja in sorted(hojas2 - hojas1):
            reporte.write(f"[HOJA FALTANTE] '{hoja}' solo existe en ARCHIVO 2\n")
            hojas_faltantes += 1

        resumen_json["resumen"]["hojas_faltantes"] = hojas_faltantes

        reporte.write("\n")

        # Comparar hojas comunes
        hojas_comparadas = 0
        for hoja in sorted(hojas1 & hojas2):
            ws1 = wb1[hoja]
            ws2 = wb2[hoja]

            encabezados1 = obtener_encabezados(ws1)
            encabezados2 = obtener_encabezados(ws2)

            reporte.write(f"HOJA: {hoja}\n")
            reporte.write("-" * 50 + "\n")

            diferencias_hoja = 0
            hoja_json = {
                "nombre": hoja,
                "diferencias": 0,
                "detalle": []
            }

            # Validación de filas y columnas
            if ws1.max_row != ws2.max_row:
                reporte.write(f"[DIFERENCIA FILAS] Archivo1={ws1.max_row} | Archivo2={ws2.max_row}\n")
                hoja_json["detalle"].append({
                    "campo": "Filas",
                    "fila": 0,
                    "archivo1": ws1.max_row,
                    "archivo2": ws2.max_row
                })
                diferencias_hoja += 1

            if ws1.max_column != ws2.max_column:
                reporte.write(f"[DIFERENCIA COLUMNAS] Archivo1={ws1.max_column} | Archivo2={ws2.max_column}\n")
                hoja_json["detalle"].append({
                    "campo": "Columnas",
                    "fila": 0,
                    "archivo1": ws1.max_column,
                    "archivo2": ws2.max_column
                })
                diferencias_hoja += 1

            max_fila = max(ws1.max_row, ws2.max_row)
            max_col = max(ws1.max_column, ws2.max_column)

            for fila in range(2, max_fila + 1):
                for col in range(1, max_col + 1):
                    v1 = ws1.cell(row=fila, column=col).value
                    v2 = ws2.cell(row=fila, column=col).value

                    if v1 != v2:
                        nombre_col = encabezados1.get(col) or encabezados2.get(col)
                        reporte.write(
                            f"{hoja}.{nombre_col} (fila {fila}): "
                            f"Archivo1='{v1}' | Archivo2='{v2}'\n"
                        )
                        total_diferencias += 1
                        diferencias_hoja += 1
                        hoja_json["detalle"].append({
                            "campo": nombre_col,
                            "fila": fila,
                            "archivo1": v1,
                            "archivo2": v2
                        })

            if diferencias_hoja == 0:
                reporte.write("Sin diferencias en esta hoja.\n")
                
                diferencias_por_hoja[hoja] = diferencias_hoja
                hoja_json["diferencias"] = diferencias_hoja
                resumen_json["hojas"].append(hoja_json)
                hojas_comparadas += 1
            else:
                diferencias_por_hoja[hoja] = diferencias_hoja
                hoja_json["diferencias"] = diferencias_hoja
                resumen_json["hojas"].append(hoja_json)
                hojas_comparadas += 1

            reporte.write("\n")

        resumen_json["resumen"]["hojas_comparadas"] = hojas_comparadas
        resumen_json["resumen"]["total_diferencias"] = total_diferencias

        # RESUMEN FINAL
        reporte.write("\n")
        reporte.write("RESUMEN DE DIFERENCIAS\n")
        reporte.write("=" * 70 + "\n")
        reporte.write(f"Hojas comparadas        : {hojas_comparadas}\n")
        reporte.write(f"Hojas faltantes         : {hojas_faltantes}\n")
        reporte.write(f"Total de diferencias    : {total_diferencias}\n\n")

        if diferencias_por_hoja:
            reporte.write("Diferencias por hoja:\n")
            for hoja, cantidad in diferencias_por_hoja.items():
                reporte.write(f" - {hoja}: {cantidad}\n")
        else:
            reporte.write("No se encontraron diferencias.\n")

    # Guardar JSON
    with open(archivo_json, "w", encoding="utf-8") as json_file:
        json.dump(resumen_json, json_file, ensure_ascii=False, indent=4, default=default_converter )

    print("Reporte de texto y JSON generados correctamente.")

if __name__ == "__main__":
    print("Selecciona el PRIMER archivo XLSX")
    archivo1 = seleccionar_archivo("Selecciona el primer XLSX")

    print("Selecciona el SEGUNDO archivo XLSX")
    archivo2 = seleccionar_archivo("Selecciona el segundo XLSX")

    print("Selecciona dónde guardar el reporte de texto")
    archivo_reporte = seleccionar_guardado()

    print("Selecciona dónde guardar el archivo JSON")
    archivo_json = seleccionar_guardado()

    if archivo1 and archivo2 and archivo_reporte and archivo_json:
        comparar_xlsx(archivo1, archivo2, archivo_reporte, archivo_json)
    else:
        print("Operación cancelada.")
